"""Bounded synthetic UpSeller-style XLSX workbook adapter."""

from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from io import BytesIO, StringIO
from time import monotonic
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree

from openpyxl import load_workbook

from src.adapters.protocol import (
    AdapterLimits,
    RecognitionResult,
    SourceShapeInvalid,
    StandardizedArtifact,
)
from src.adapters.validation import (
    canonicalize_role_records,
    role_has_required_fields,
    validate_mapping,
)
from src.services.canonical_contracts import RowOrigin
from src.synthetic.boundary import (
    validate_safe_import_records,
    validate_synthetic_records,
)

MAX_ZIP_ENTRIES = 128
MAX_EXPANDED_BYTES = 32 * 1024 * 1024
MAX_SHEETS = 12
MAX_ROWS_PER_SHEET = 20_000
MAX_COLUMNS = 64
MAX_CELLS = 250_000
MAX_CELL_CHARS = 4_096
MAX_PARSE_SECONDS = 5.0
OPENXML_STRUCTURAL_URIS = frozenset(
    {
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
        "extended-properties",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
        "officeDocument",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
        "sharedStrings",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
        "worksheet",
        "http://schemas.openxmlformats.org/package/2006/relationships/"
        "metadata/core-properties",
    }
)
SHEET_ROLES = {
    "Daily Sales": "daily_sales",
    "Advertising": "shopee_advertising",
    "Inventory": "product_inventory_sales",
    "Receipt Lots": "inventory_receipt_lot",
    "Expenses": "operating_expense",
}
ALLOWED_FIELDS = {
    "date",
    "order_id",
    "store_id",
    "sku_id",
    "units",
    "unit_price_brl",
    "discount_brl",
    "gross_sales_brl",
    "net_sales_brl",
    "spend_brl",
    "impressions",
    "clicks",
    "attributed_orders",
    "on_hand_units",
    "inbound_units",
    "lot_id",
    "receipt_date",
    "quantity_received",
    "unit_cost_brl",
    "expense_id",
    "period_start",
    "period_end",
    "expense_type",
    "amount_brl",
    "product_name",
    "category",
    "movement_id",
    "movement_type",
    "quantity",
    "outbound_id",
    "settlement_id",
    "settlement_date",
    "fee_id",
    "fee_rate",
    "fee_brl",
    "payout_brl",
    "fx_id",
    "currency",
    "brl_per_usd",
    "policy_id",
    "lead_time_days",
    "safety_stock_units",
    "reorder_point_units",
    "target_cover_days",
    "scenario_id",
    "scope",
    "source_classification",
}
CSV_SOURCE_PROFILES = {
    "daily_sales": frozenset(
        {"date", "order_id", "store_id", "sku_id", "units", "gross_sales_brl"}
    ),
    "product_sales": frozenset(
        {"sku_id", "product_name", "category", "unit_price_brl"}
    ),
    "product_inventory_sales": frozenset(
        {"date", "store_id", "sku_id", "on_hand_units", "inbound_units"}
    ),
    "inventory_movement": frozenset(
        {"movement_id", "date", "store_id", "sku_id", "movement_type", "quantity"}
    ),
    "inventory_receipt_lot": frozenset(
        {
            "lot_id",
            "receipt_date",
            "store_id",
            "sku_id",
            "quantity_received",
            "unit_cost_brl",
        }
    ),
    "outbound_event": frozenset(
        {"outbound_id", "date", "store_id", "sku_id", "quantity"}
    ),
    "operating_expense": frozenset(
        {"expense_id", "period_start", "period_end", "expense_type", "amount_brl"}
    ),
    "settlement": frozenset(
        {"fee_id", "period_start", "period_end", "store_id", "fee_brl"}
    ),
    "marketplace_settlement": frozenset(
        {"settlement_id", "period_start", "period_end", "store_id", "payout_brl"}
    ),
    "fx_assumption": frozenset(
        {"fx_id", "period_start", "period_end", "currency", "brl_per_usd"}
    ),
    "replenishment_policy": frozenset(
        {
            "policy_id",
            "sku_id",
            "lead_time_days",
            "safety_stock_units",
            "reorder_point_units",
            "target_cover_days",
        }
    ),
}


class UpsellerExcelAdapter:
    adapter_id = "upseller_excel"
    adapter_version = "1.0.0"
    canonical_schema = "canonical.import.v1"
    limits = AdapterLimits(
        max_rows=MAX_ROWS_PER_SHEET,
        max_columns=MAX_COLUMNS,
        max_cell_chars=MAX_CELL_CHARS,
        max_parse_seconds=MAX_PARSE_SECONDS,
        max_expanded_bytes=MAX_EXPANDED_BYTES,
        max_sheets=MAX_SHEETS,
    )

    def recognizes(
        self,
        filename: str,
        media_type: str,
        content: bytes,
        source_kind: str = "legacy_synthetic",
    ) -> bool:
        return (
            filename.lower().endswith(".xlsx")
            and media_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            and content.startswith(b"PK")
        )

    def recognize(
        self,
        content: bytes,
        source_kind: str = "legacy_synthetic",
    ) -> RecognitionResult:
        tables = _parse(content, source_kind)
        detected_roles = tuple(SHEET_ROLES[name] for name in tables)
        fields = tuple(sorted({field for rows in tables.values() for field in rows[0]}))
        return RecognitionResult(
            adapter_id=self.adapter_id,
            adapter_version=self.adapter_version,
            source_role=detected_roles[0],
            record_count=sum(len(rows) for rows in tables.values()),
            source_fields=fields,
            suggested_mapping={field: field for field in fields},
            details={
                "detected_roles": list(detected_roles),
                "canonical_schema": self.canonical_schema,
            },
        )

    def standardize(
        self,
        content: bytes,
        mapping: dict[str, str],
        source_kind: str = "legacy_synthetic",
        *,
        source_name: str = "source",
    ) -> StandardizedArtifact:
        tables = _parse(content, source_kind)
        fields = {field for rows in tables.values() for field in rows[0]}
        roles = tuple(SHEET_ROLES[sheet] for sheet in tables)
        validate_mapping(
            fields,
            mapping,
            allowed_fields=ALLOWED_FIELDS,
            roles=roles,
        )
        canonical_tables = {
            SHEET_ROLES[sheet]: [
                {mapping[field]: value for field, value in row.items()}
                for row in rows
            ]
            for sheet, rows in tables.items()
        }
        canonical_tables = {
            role: canonicalize_role_records(role, records)
            for role, records in canonical_tables.items()
        }
        if "operating_expense" in canonical_tables:
            canonical_tables["operating_expense"] = _scope_operating_expenses(
                canonical_tables["operating_expense"]
            )
        row_provenance = {
            SHEET_ROLES[sheet]: [
                RowOrigin(
                    source_kind="upload",
                    source_name=source_name,
                    sheet_name=sheet,
                    row_number=row_number,
                ).safe_projection()
                for row_number in range(2, len(rows) + 2)
            ]
            for sheet, rows in tables.items()
        }
        serialized = (
            json.dumps(
                {
                    "schema_version": "canonical.import.v1",
                    "source_classification": (
                        "pure_synthetic"
                        if source_kind == "legacy_synthetic"
                        else "operator_upload"
                    ),
                    "tables": canonical_tables,
                    "row_provenance": row_provenance,
                },
                sort_keys=True,
                separators=(",", ":"),
            )
            + "\n"
        ).encode()
        preview = tuple(
            record
            for records in canonical_tables.values()
            for record in records
        )[:25]
        return StandardizedArtifact(
            content=serialized,
            record_count=sum(len(rows) for rows in canonical_tables.values()),
            preview_records=preview,
            quality_report={
                "status": "passed",
                "record_count": sum(len(rows) for rows in canonical_tables.values()),
                "table_roles": list(canonical_tables),
            },
        )


def validate_safe_xlsx_package(content: bytes) -> None:
    """Validate the passive OpenXML boundary without assigning an import role."""

    _validate_zip(content)


def _parse(
    content: bytes,
    source_kind: str = "legacy_synthetic",
) -> dict[str, list[dict[str, object]]]:
    started = monotonic()
    _validate_zip(content)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (OSError, ValueError, BadZipFile) as error:
        raise SourceShapeInvalid("xlsx_parse_failed") from error
    try:
        if not 1 <= len(workbook.worksheets) <= MAX_SHEETS:
            raise SourceShapeInvalid("xlsx_sheet_limit_exceeded")
        tables = {}
        total_cells = 0
        for worksheet in workbook.worksheets:
            if worksheet.title not in SHEET_ROLES:
                raise SourceShapeInvalid("xlsx_sheet_not_supported")
            iterator = worksheet.iter_rows()
            try:
                headers = tuple(str(cell.value or "").strip() for cell in next(iterator))
            except StopIteration as error:
                raise SourceShapeInvalid("xlsx_sheet_empty") from error
            if (
                not headers
                or len(headers) > MAX_COLUMNS
                or len(set(headers)) != len(headers)
                or not set(headers) <= ALLOWED_FIELDS
                or (
                    source_kind == "legacy_synthetic"
                    and "source_classification" not in headers
                )
                or not role_has_required_fields(
                    SHEET_ROLES[worksheet.title], set(headers)
                )
            ):
                raise SourceShapeInvalid("xlsx_headers_invalid")
            records = []
            for row in iterator:
                if len(records) >= MAX_ROWS_PER_SHEET:
                    raise SourceShapeInvalid("xlsx_row_limit_exceeded")
                total_cells += len(row)
                if total_cells > MAX_CELLS or monotonic() - started > MAX_PARSE_SECONDS:
                    raise SourceShapeInvalid("xlsx_parse_budget_exceeded")
                record = {}
                for header, cell in zip(headers, row, strict=True):
                    if getattr(cell, "data_type", None) == "f":
                        raise SourceShapeInvalid("xlsx_formula_not_allowed")
                    value = _normalized_cell_value(cell.value)
                    if value is not None and len(str(value)) > MAX_CELL_CHARS:
                        raise SourceShapeInvalid("xlsx_cell_limit_exceeded")
                    record[header] = "" if value is None else value
                records.append(record)
            if not records:
                raise SourceShapeInvalid("xlsx_sheet_has_no_records")
            validator = (
                validate_synthetic_records
                if source_kind == "legacy_synthetic"
                else validate_safe_import_records
            )
            validator(records)
            tables[worksheet.title] = records
        return tables
    finally:
        workbook.close()


def _validate_zip(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ZIP_ENTRIES:
                raise SourceShapeInvalid("xlsx_zip_entry_limit_exceeded")
            if sum(entry.file_size for entry in entries) > MAX_EXPANDED_BYTES:
                raise SourceShapeInvalid("xlsx_expanded_size_limit_exceeded")
            names = {
                entry.filename.lower(): entry.filename
                for entry in entries
            }
            if len(names) != len(entries):
                raise SourceShapeInvalid("xlsx_zip_entry_name_collision")
            allowed_names = {
                "[content_types].xml",
                "_rels/.rels",
                "docprops/app.xml",
                "docprops/core.xml",
                "xl/_rels/workbook.xml.rels",
                "xl/sharedstrings.xml",
                "xl/styles.xml",
                "xl/workbook.xml",
            }
            if any(
                name not in allowed_names
                and not name.startswith("xl/theme/")
                and not (
                    name.startswith("xl/worksheets/")
                    and name.endswith(".xml")
                )
                for name in names
            ):
                raise SourceShapeInvalid("xlsx_package_part_not_allowed")
            for name, original_name in names.items():
                if name.endswith((".xml", ".rels")):
                    _validate_xml_content(archive.read(original_name))
            forbidden_fragments = (
                "/activex/",
                "/comments",
                "/connections",
                "/ctrlprops/",
                "/customxml/",
                "/drawings/",
                "/embeddings/",
                "/externallinks/",
                "/media/",
                "/persons/",
                "/printersettings/",
                "/threadedcomments/",
                "/worksheets/_rels/",
            )
            if any(
                name.endswith(("vbaproject.bin", ".exe", ".js", ".bin", ".vml"))
                or any(fragment in f"/{name}" for fragment in forbidden_fragments)
                or name == "docprops/custom.xml"
                for name in names
            ):
                raise SourceShapeInvalid("xlsx_active_content_not_allowed")
            for name in names:
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    xml = archive.read(names[name]).lower()
                    if any(
                        marker in xml
                        for marker in (
                            b"<hyperlink",
                            b"<headerfooter",
                            b"<legacydrawing",
                            b"<oleobject",
                            b"<drawing",
                        )
                    ):
                        raise SourceShapeInvalid("xlsx_active_content_not_allowed")
            for relationship_name in (
                "_rels/.rels",
                "xl/_rels/workbook.xml.rels",
            ):
                if (
                    relationship_name in names
                    and b'targetmode="external"'
                    in archive.read(names[relationship_name]).lower()
                ):
                    raise SourceShapeInvalid("xlsx_active_content_not_allowed")
            if "xl/workbook.xml" in names:
                workbook_xml = archive.read(names["xl/workbook.xml"]).lower()
                if b"<definedname" in workbook_xml:
                    raise SourceShapeInvalid("xlsx_defined_name_not_allowed")
    except BadZipFile as error:
        raise SourceShapeInvalid("xlsx_zip_invalid") from error


def _validate_xml_content(content: bytes) -> None:
    try:
        parser = ElementTree.XMLParser(
            target=ElementTree.TreeBuilder(insert_comments=True)
        )
        root = ElementTree.fromstring(content, parser=parser)
    except ElementTree.ParseError as error:
        raise SourceShapeInvalid("xlsx_xml_invalid") from error
    records = []
    records.extend(
        {"xlsx_package_comment": comment.decode("utf-8", errors="strict")}
        for comment in re.findall(rb"<!--(.*?)-->", content, flags=re.DOTALL)
    )
    for element in root.iter():
        values = []
        if element.text and element.text.strip():
            values.append(element.text)
        if element.tail and element.tail.strip():
            values.append(element.tail)
        values.extend(
            value
            for value in element.attrib.values()
            if value
            and value.strip()
            and value not in OPENXML_STRUCTURAL_URIS
        )
        records.extend(
            {"xlsx_package_value": value}
            for value in values
        )
    validate_synthetic_records(records)


def _normalized_cell_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return (
            value.date().isoformat()
            if value.time() == datetime.min.time()
            else value.isoformat()
        )
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (str, int, float)) and not isinstance(value, bool):
        return value
    raise SourceShapeInvalid("xlsx_cell_type_not_allowed")


class UpsellerCsvAdapter:
    """Recognize the declared pure-synthetic UpSeller export roles by shape."""

    adapter_id = "upseller_csv"
    adapter_version = "1.0.0"
    canonical_schema = "canonical.import.v1"
    limits = AdapterLimits(
        max_rows=MAX_ROWS_PER_SHEET,
        max_columns=MAX_COLUMNS,
        max_cell_chars=MAX_CELL_CHARS,
        max_parse_seconds=MAX_PARSE_SECONDS,
    )

    def recognizes(
        self,
        filename: str,
        media_type: str,
        content: bytes,
        source_kind: str = "legacy_synthetic",
    ) -> bool:
        if not filename.lower().endswith(".csv") or media_type != "text/csv":
            return False
        try:
            headers, _records = _parse_upseller_csv(content, source_kind)
            _detect_csv_role(headers)
        except SourceShapeInvalid:
            return False
        return True

    def recognize(
        self,
        content: bytes,
        source_kind: str = "legacy_synthetic",
    ) -> RecognitionResult:
        headers, records = _parse_upseller_csv(content, source_kind)
        source_role = _detect_csv_role(headers)
        return RecognitionResult(
            adapter_id=self.adapter_id,
            adapter_version=self.adapter_version,
            source_role=source_role,
            record_count=len(records),
            source_fields=headers,
            suggested_mapping={field: field for field in headers},
            details={"delimiter": ",", "canonical_schema": self.canonical_schema},
        )

    def standardize(
        self,
        content: bytes,
        mapping: dict[str, str],
        source_kind: str = "legacy_synthetic",
        *,
        source_name: str = "source",
    ) -> StandardizedArtifact:
        headers, records = _parse_upseller_csv(content, source_kind)
        source_role = _detect_csv_role(headers)
        validate_mapping(
            set(headers),
            mapping,
            allowed_fields=ALLOWED_FIELDS,
            roles=(source_role,),
        )
        canonical = [
            {mapping[field]: value for field, value in record.items()}
            for record in records
        ]
        canonical = canonicalize_role_records(source_role, canonical)
        if source_role == "operating_expense":
            canonical = _scope_operating_expenses(canonical)
        row_provenance = [
            RowOrigin(
                source_kind="upload",
                source_name=source_name,
                sheet_name=None,
                row_number=row_number,
            ).safe_projection()
            for row_number in range(2, len(canonical) + 2)
        ]
        serialized = (
            json.dumps(
                {
                    "schema_version": self.canonical_schema,
                    "source_classification": (
                        "pure_synthetic"
                        if source_kind == "legacy_synthetic"
                        else "operator_upload"
                    ),
                    "tables": {source_role: canonical},
                    "row_provenance": {source_role: row_provenance},
                },
                sort_keys=True,
                separators=(",", ":"),
            )
            + "\n"
        ).encode()
        return StandardizedArtifact(
            content=serialized,
            record_count=len(canonical),
            preview_records=tuple(canonical[:25]),
            quality_report={
                "status": "passed",
                "record_count": len(canonical),
                "table_roles": [source_role],
            },
        )


def _scope_operating_expenses(
    records: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Normalize legacy shared expenses without erasing explicit store scope."""

    scoped_records: list[dict[str, object]] = []
    for record in records:
        store_id = str(record.get("store_id") or "").strip()
        declared_scope = str(record.get("scope") or "").strip().lower()
        if store_id:
            if declared_scope not in {"", "store"}:
                raise SourceShapeInvalid("operating_expense_scope_invalid")
            scope = "store"
        else:
            if declared_scope not in {"", "shared"}:
                raise SourceShapeInvalid("operating_expense_scope_invalid")
            scope = "shared"
        scoped_records.append({**record, "scope": scope})
    return scoped_records


def _parse_upseller_csv(
    content: bytes,
    source_kind: str = "legacy_synthetic",
) -> tuple[tuple[str, ...], list[dict[str, str]]]:
    started = monotonic()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise SourceShapeInvalid("csv_must_be_utf8") from error
    reader = csv.DictReader(StringIO(text, newline=""))
    headers = tuple(reader.fieldnames or ())
    if (
        not headers
        or len(headers) > MAX_COLUMNS
        or len(set(headers)) != len(headers)
        or not set(headers) <= ALLOWED_FIELDS
        or (
            source_kind == "legacy_synthetic"
            and "source_classification" not in headers
        )
    ):
        raise SourceShapeInvalid("csv_headers_invalid")
    records = []
    for record in reader:
        if len(records) >= MAX_ROWS_PER_SHEET:
            raise SourceShapeInvalid("csv_row_limit_exceeded")
        if monotonic() - started > MAX_PARSE_SECONDS:
            raise SourceShapeInvalid("csv_parse_budget_exceeded")
        if None in record or any(record.get(field) is None for field in headers):
            raise SourceShapeInvalid("csv_column_count_mismatch")
        normalized = {field: record.get(field, "") for field in headers}
        if any(len(value or "") > MAX_CELL_CHARS for value in normalized.values()):
            raise SourceShapeInvalid("csv_cell_limit_exceeded")
        records.append(normalized)
    if not records:
        raise SourceShapeInvalid("csv_has_no_records")
    validator = (
        validate_synthetic_records
        if source_kind == "legacy_synthetic"
        else validate_safe_import_records
    )
    validator(records)
    return headers, records


def _detect_csv_role(headers: tuple[str, ...]) -> str:
    fields = set(headers)
    matches = [
        (role, required)
        for role, required in CSV_SOURCE_PROFILES.items()
        if required <= fields
    ]
    if len(matches) != 1:
        raise SourceShapeInvalid("upseller_source_role_ambiguous_or_unknown")
    role, _required = matches[0]
    return "settlement" if role == "marketplace_settlement" else role
