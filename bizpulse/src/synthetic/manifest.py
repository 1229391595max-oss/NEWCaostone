"""Load and verify a declared pure-synthetic fixture bundle."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import date
from hashlib import sha256
from pathlib import Path, PurePosixPath
from typing import Any, Literal, cast

from openpyxl import load_workbook

from src.services.canonical_contracts import StoreDescriptor
from src.synthetic.contracts import SyntheticBundle, SyntheticFile, SyntheticManifest

ALLOWED_MEDIA_TYPES = {
    ".csv": "text/csv",
    ".json": "application/json",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
EMAIL_PATTERN = re.compile(r"\b[^\s@]+@[^\s@]+\.[^\s@]+\b", re.IGNORECASE)
PHONE_PATTERN = re.compile(r"(?:\+\d{8,15}|\b\d{10,15}\b)")
CREDENTIAL_PATTERN = re.compile(
    r"(?:sk-(?:proj-)?[A-Za-z0-9_-]{12,}|AccountKey=|postgres(?:ql)?://|"
    r"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----)",
    re.IGNORECASE,
)
URL_PATTERN = re.compile(r"https?://", re.IGNORECASE)
ADDRESS_PATTERN = re.compile(
    r"\b(?:street|avenue|boulevard|road|rua|endereco|address)\b",
    re.IGNORECASE,
)
FORBIDDEN_SOURCE_PATTERN = re.compile(
    r"\b(?:google_trends|mercado_live|real|private)\b",
    re.IGNORECASE,
)


@dataclass(frozen=True, slots=True)
class BoundaryViolation:
    file: str
    field: str
    rule: str

    def __str__(self) -> str:
        return f"{self.file}:{self.field}:{self.rule}"


def verify_bundle_directory(root: Path) -> tuple[BoundaryViolation, ...]:
    violations: list[BoundaryViolation] = []
    try:
        manifest_payload = json.loads((root / "manifest.json").read_text())
        declarations = _declarations(manifest_payload)
        scenario_ids = set(manifest_payload["scenario_ids"])
        _store_catalog(manifest_payload)
    except (OSError, KeyError, TypeError, ValueError, json.JSONDecodeError):
        return (BoundaryViolation("manifest.json", "manifest", "manifest_invalid"),)

    if manifest_payload.get("source_classification") != "pure_synthetic":
        violations.append(
            BoundaryViolation(
                "manifest.json",
                "source_classification",
                "source_classification_invalid",
            )
        )
    if manifest_payload.get("currency") != "BRL":
        violations.append(
            BoundaryViolation("manifest.json", "currency", "currency_invalid")
        )
    declared_paths = set(declarations)
    actual_paths = {
        path.relative_to(root).as_posix()
        for path in root.rglob("*")
        if path.is_file() and path.name != "manifest.json"
    }
    for path in sorted(actual_paths - declared_paths):
        violations.append(BoundaryViolation(path, "file", "undeclared_file"))
    for path in sorted(declared_paths - actual_paths):
        violations.append(BoundaryViolation(path, "file", "declared_file_missing"))

    for relative_path in sorted(actual_paths):
        candidate = root / relative_path
        declaration = declarations.get(relative_path)
        if declaration is not None:
            content = candidate.read_bytes()
            if sha256(content).hexdigest() != declaration["sha256"]:
                violations.append(
                    BoundaryViolation(relative_path, "file", "sha256_mismatch")
                )
            expected_media = ALLOWED_MEDIA_TYPES.get(candidate.suffix.lower())
            if expected_media is None or declaration["media_type"] != expected_media:
                violations.append(
                    BoundaryViolation(relative_path, "file", "media_type_invalid")
                )
        try:
            if candidate.suffix.lower() == ".csv":
                rows, row_count = _csv_values(candidate, violations)
            elif candidate.suffix.lower() == ".json":
                rows, row_count = _json_values(candidate, violations)
            elif candidate.suffix.lower() == ".xlsx":
                rows, row_count = _xlsx_values(candidate, violations)
            else:
                violations.append(
                    BoundaryViolation(relative_path, "file", "file_type_invalid")
                )
                continue
        except (OSError, UnicodeError, csv.Error, ValueError):
            violations.append(BoundaryViolation(relative_path, "file", "parse_failed"))
            continue
        if declaration is not None and row_count != int(declaration["row_count"]):
            violations.append(BoundaryViolation(relative_path, "file", "row_count_mismatch"))
        for field, value in rows:
            violations.extend(
                _value_violations(relative_path, field, value, scenario_ids)
            )
    return tuple(violations)


def load_bundle(root: Path) -> SyntheticBundle:
    violations = verify_bundle_directory(root)
    if violations:
        rules = ",".join(sorted({violation.rule for violation in violations}))
        raise ValueError(f"synthetic_bundle_invalid:{rules}")
    payload = json.loads((root / "manifest.json").read_text())
    declarations = _declarations(payload)
    files = tuple(
        SyntheticFile(
            relative_path=path,
            content=(root / path).read_bytes(),
            media_type=str(declaration["media_type"]),
            row_count=int(declaration["row_count"]),
        )
        for path, declaration in sorted(declarations.items())
    )
    manifest = SyntheticManifest(
        schema_version=str(payload["schema_version"]),
        generator_version=str(payload["generator"]["version"]),
        generator_source_sha256=str(payload["generator"]["source_sha256"]),
        seed=int(payload["seed"]),
        source_classification=str(payload["source_classification"]),
        currency=str(payload["currency"]),
        date_range=(
            str(payload["date_range"]["start"]),
            str(payload["date_range"]["end"]),
        ),
        reporting_period=(
            str(payload.get("reporting_period", payload["date_range"])["start"]),
            str(payload.get("reporting_period", payload["date_range"])["end"]),
        ),
        scenario_ids=tuple(str(item) for item in payload["scenario_ids"]),
        store_catalog=_store_catalog(payload),
        files=files,
    )
    return SyntheticBundle(manifest, (root / "manifest.json").read_bytes(), files)


def _store_catalog(payload: dict[str, Any]) -> tuple[StoreDescriptor, ...]:
    raw_catalog = payload.get("store_catalog")
    if not isinstance(raw_catalog, list) or not raw_catalog:
        raise ValueError("manifest_store_catalog_invalid")
    catalog = []
    seen = set()
    for raw in raw_catalog:
        if not isinstance(raw, dict):
            raise ValueError("manifest_store_catalog_invalid")
        try:
            store_id = str(raw["store_id"])
            opened_on = raw.get("opened_on")
            lifecycle = str(raw["lifecycle"])
            if (
                not store_id
                or store_id in seen
                or lifecycle not in {"established", "new"}
                or type(raw["has_data"]) is not bool
            ):
                raise ValueError("manifest_store_catalog_invalid")
            descriptor = StoreDescriptor(
                store_id=store_id,
                display_name_en=str(raw["display_name_en"]),
                display_name_zh=str(raw["display_name_zh"]),
                currency=str(raw["currency"]),
                opened_on=date.fromisoformat(opened_on) if opened_on else None,
                lifecycle=cast(Literal["established", "new"], lifecycle),
                has_data=raw["has_data"],
            )
        except (KeyError, TypeError, ValueError) as error:
            raise ValueError("manifest_store_catalog_invalid") from error
        if not (
            descriptor.display_name_en
            and descriptor.display_name_zh
            and descriptor.currency
        ):
            raise ValueError("manifest_store_catalog_invalid")
        seen.add(store_id)
        catalog.append(descriptor)
    return tuple(catalog)


def _declarations(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    declarations = {}
    for raw in payload["files"]:
        relative_path = str(raw["path"])
        normalized = PurePosixPath(relative_path)
        if (
            normalized.is_absolute()
            or ".." in normalized.parts
            or str(normalized) != relative_path
            or relative_path in declarations
        ):
            raise ValueError("manifest_path_invalid")
        media_type = str(raw["media_type"])
        digest = str(raw["sha256"])
        row_count = int(raw["row_count"])
        if (
            ALLOWED_MEDIA_TYPES.get(normalized.suffix.lower()) != media_type
            or re.fullmatch(r"[0-9a-f]{64}", digest) is None
            or row_count < 0
        ):
            raise ValueError("manifest_declaration_invalid")
        declarations[relative_path] = raw
    return declarations


def _csv_values(
    path: Path,
    violations: list[BoundaryViolation],
) -> tuple[list[tuple[str, object]], int]:
    values = []
    row_count = 0
    with path.open(newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream)
        if "source_classification" not in (reader.fieldnames or ()):
            violations.append(
                BoundaryViolation(
                    path.name,
                    "source_classification",
                    "source_classification_missing",
                )
            )
        for row_index, row in enumerate(reader, start=2):
            row_count += 1
            for field, value in row.items():
                values.append((f"row{row_index}.{field or 'unknown'}", value))
    return values, row_count


def _xlsx_values(
    path: Path,
    violations: list[BoundaryViolation],
) -> tuple[list[tuple[str, object]], int]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    values = []
    row_count = 0
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows()
            headers = tuple(str(cell.value or "") for cell in next(rows))
            if "source_classification" not in headers:
                violations.append(
                    BoundaryViolation(
                        path.name,
                        f"{worksheet.title}.source_classification",
                        "source_classification_missing",
                    )
                )
            for row_index, row in enumerate(rows, start=2):
                row_count += 1
                for column_index, (header, cell) in enumerate(
                    zip(headers, row, strict=True),
                    start=1,
                ):
                    coordinate = getattr(cell, "coordinate", None) or (
                        f"R{row_index}C{column_index}"
                    )
                    field = f"{worksheet.title}.{coordinate}.{header}"
                    if cell.data_type == "f":
                        violations.append(
                            BoundaryViolation(path.name, field, "formula_cell")
                        )
                    values.append((field, cell.value))
    finally:
        workbook.close()
    return values, row_count


def _json_values(
    path: Path,
    violations: list[BoundaryViolation],
) -> tuple[list[tuple[str, object]], int]:
    payload = json.loads(path.read_text())
    if (
        not isinstance(payload, dict)
        or payload.get("schema_version") != "canonical.analysis.v1"
        or not isinstance(payload.get("tables"), dict)
    ):
        raise ValueError("analysis_bundle_invalid")
    values: list[tuple[str, object]] = []
    row_count = 0
    for role, records in sorted(payload["tables"].items()):
        if not isinstance(role, str) or not isinstance(records, list):
            raise ValueError("analysis_bundle_invalid")
        for row_index, record in enumerate(records, start=1):
            if not isinstance(record, dict):
                raise ValueError("analysis_bundle_invalid")
            row_count += 1
            if record.get("source_classification") != "pure_synthetic":
                violations.append(
                    BoundaryViolation(
                        path.name,
                        f"{role}.row{row_index}.source_classification",
                        "source_classification_invalid",
                    )
                )
            for field, value in record.items():
                values.append((f"{role}.row{row_index}.{field}", value))
    return values, row_count


def _value_violations(
    file: str,
    field: str,
    raw_value: object,
    scenario_ids: set[str],
) -> list[BoundaryViolation]:
    if raw_value is None:
        return []
    value = str(raw_value)
    rules = []
    for pattern, rule in (
        (EMAIL_PATTERN, "email_pattern"),
        (PHONE_PATTERN, "phone_pattern"),
        (CREDENTIAL_PATTERN, "credential_pattern"),
        (URL_PATTERN, "external_url"),
        (ADDRESS_PATTERN, "address_pattern"),
        (FORBIDDEN_SOURCE_PATTERN, "forbidden_source_label"),
    ):
        if pattern.search(value):
            rules.append(rule)
    field_name = field.rsplit(".", 1)[-1].lower()
    if field_name == "source_classification" and value != "pure_synthetic":
        rules.append("source_classification_invalid")
    if field_name.endswith("_id") and value:
        if field_name == "scenario_id":
            if value not in scenario_ids:
                rules.append("unapproved_identifier")
        elif not value.startswith("SYNTH-"):
            rules.append("unapproved_identifier")
    if field_name.endswith("_ids") and value:
        if any(not item.startswith("SYNTH-") for item in value.split("|")):
            rules.append("unapproved_identifier")
    return [BoundaryViolation(file, field, rule) for rule in rules]
