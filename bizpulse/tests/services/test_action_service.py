from __future__ import annotations

from dataclasses import replace
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
import pytest
from sqlalchemy import Engine, func, select

from src.actions.contracts import ActionAdjustment, ActionSource, FactRef
from src.actions.state_machine import ActionTransitionInvalid
from src.db.schema import (
    action_card_revisions,
    action_cards,
    action_decisions,
    action_exports,
    action_outcomes,
    analysis_artifacts,
    storage_objects,
)
from src.db.unit_of_work import PostgresUnitOfWork
from src.forecast.contracts import ForecastRequest, ProductCandidate
from src.repositories.operators import OperatorRepository
from src.services.action_service import (
    ActionIdempotencyConflict,
    ActionInvalid,
    ActionRevisionConflict,
    ActionService,
)
from src.services.analysis_service import AnalysisService
from src.services.forecast_service import ForecastService
from src.services.profit_bridge_service import ProfitBridgeService
from src.services.public_release_service import PUBLIC_ANALYSIS_SCOPE
from src.storage.lifecycle import StorageLifecycle
from src.synthetic.generator import generate_demo
from src.synthetic.seed import seed_demo
from tests.import_support import MemoryWorkflowStorage, WORKSPACE_ID

NOW = datetime(2026, 8, 14, 15, 0, tzinfo=UTC)
MAIN_ACTION_SCOPE = {"currency": "BRL", "store_id": "SYNTH-STORE-01"}


class PromoteAcknowledgementLostStorage(MemoryWorkflowStorage):
    def __init__(self) -> None:
        super().__init__()
        self.failures_remaining = 0

    def promote(self, staged_key, final_key, expected_sha256):
        promoted = super().promote(staged_key, final_key, expected_sha256)
        if self.failures_remaining:
            self.failures_remaining -= 1
            raise RuntimeError("injected_export_promote_acknowledgement_lost")
        return promoted


class StagingDeleteFailureStorage(MemoryWorkflowStorage):
    armed = False

    def delete(self, key, *, expected_etag=None) -> None:
        if self.armed and "/staging/" in key:
            self.armed = False
            raise RuntimeError("injected_action_staging_delete_failure")
        super().delete(key, expected_etag=expected_etag)


class UnrecoverablePromotionStorage(MemoryWorkflowStorage):
    armed = False

    def promote(self, staged_key, final_key, expected_sha256):
        promoted = super().promote(staged_key, final_key, expected_sha256)
        if self.armed:
            raise RuntimeError("injected_unrecoverable_promotion")
        return promoted

    def inventory(self, prefix):
        if self.armed:
            raise RuntimeError("injected_promotion_inventory_unavailable")
        return super().inventory(prefix)

    def open_verified(self, key, expected_sha256, max_bytes):
        if self.armed and "/exports/" in key:
            raise RuntimeError("injected_promotion_read_unavailable")
        return super().open_verified(key, expected_sha256, max_bytes)


class FirstCommitFailsUnitOfWork(PostgresUnitOfWork):
    commit_calls = 0

    def commit(self) -> None:
        type(self).commit_calls += 1
        if type(self).commit_calls == 1:
            self.rollback()
            raise RuntimeError("injected_action_ledger_commit_failure")
        super().commit()


def _source(
    dataset_version_id,
    analysis_run_id,
    facts: tuple[FactRef, ...],
    *,
    suggestion: str = "Reorder 40 units",
) -> ActionSource:
    quantity_fact = next(
        item for item in facts if item.alias.endswith(".recommended_quantity")
    )
    date_fact = next(item for item in facts if item.alias.endswith(".latest_order_date"))
    priority_fact = next(item for item in facts if item.alias.endswith(".priority"))
    cash_fact = next(
        (item for item in facts if item.alias.endswith(".cash_required.value")),
        None,
    )
    assert quantity_fact.value is not None
    assert date_fact.value is not None
    assert priority_fact.value is not None
    target = quantity_fact.alias.split("|", 1)[0].split(":", 1)[1]
    return ActionSource(
        source_type="deterministic_rule",
        dataset_version_id=dataset_version_id,
        suggestion=suggestion,
        target=target,
        period_start=date.fromisoformat(str(PUBLIC_ANALYSIS_SCOPE["period_start"])),
        period_end=date.fromisoformat(str(PUBLIC_ANALYSIS_SCOPE["period_end"])),
        scope=dict(PUBLIC_ANALYSIS_SCOPE),
        quantity=Decimal(quantity_fact.value),
        budget_brl=(
            Decimal(cash_fact.value)
            if cash_fact is not None and cash_fact.value is not None
            else None
        ),
        action_date=date.fromisoformat(date_fact.value),
        threshold=None,
        expected_impact={"priority": priority_fact.value},
        confidence="medium",
        limitations=("synthetic_demo_only",),
        analysis_run_id=analysis_run_id,
        forecast_id=None,
        bridge_id=None,
        chat_turn_id=None,
        chat_tool=None,
        answer_version=None,
    )


def _action_authority(
    engine: Engine,
    storage: MemoryWorkflowStorage,
    dataset_version_id,
) -> tuple[object, tuple[FactRef, ...]]:
    analyses = AnalysisService(engine, storage, WORKSPACE_ID, clock=lambda: NOW)
    plan = analyses.plan(
        "replenishment",
        dataset_version_id,
        PUBLIC_ANALYSIS_SCOPE,
    )
    run = analyses.run(plan, "action-authority-replenishment")
    evidence = {item.alias: item for item in analyses.get_evidence(run.run_id)}
    snapshot = analyses.get_snapshot(run.run_id)
    result_item = snapshot["result"]["items"][0]
    base_alias = f"replenishment:{result_item['sku_id']}"
    authority = evidence[base_alias]
    fact_values = (
        ("recommended_quantity", result_item["recommended_quantity"]),
        ("latest_order_date", result_item["latest_order_date"]),
        ("priority", result_item["priority"]),
        ("cash_required.value", result_item["cash_required"]["value"]),
    )
    facts = tuple(
        FactRef(
            alias=f"{base_alias}|result.items.{result_item['sku_id']}.{field}",
            evidence_state=authority.evidence_state,
            source_ref=(
                f"analysis:{run.run_id}:{base_alias}|"
                f"result.items.{result_item['sku_id']}.{field}"
            ),
            value=str(value) if value is not None else None,
        )
        for field, value in fact_values
    )
    return run.run_id, facts


def _service(
    migrated_engine: Engine,
    storage: MemoryWorkflowStorage | None = None,
):
    storage = storage or MemoryWorkflowStorage()
    with PostgresUnitOfWork(migrated_engine) as uow:
        OperatorRepository(uow.connection).create_workspace(WORKSPACE_ID)
    seeded = seed_demo(
        generate_demo(),
        PostgresUnitOfWork(migrated_engine),
        storage,
        now=NOW,
    )
    analysis_run_id, facts = _action_authority(
        migrated_engine,
        storage,
        seeded.dataset_version_id,
    )
    return (
        ActionService(
            migrated_engine,
            storage,
            WORKSPACE_ID,
            clock=lambda: NOW,
        ),
        storage,
        seeded.dataset_version_id,
        analysis_run_id,
        facts,
    )


def test_operator_action_lifecycle_is_append_only_and_idempotent(
    migrated_engine: Engine,
) -> None:
    service, storage, version_id, analysis_run_id, facts = _service(migrated_engine)
    with migrated_engine.connect() as connection:
        baseline_counts = {
            "cards": connection.scalar(select(func.count()).select_from(action_cards)),
            "revisions": connection.scalar(
                select(func.count()).select_from(action_card_revisions)
            ),
            "decisions": connection.scalar(
                select(func.count()).select_from(action_decisions)
            ),
            "exports": connection.scalar(select(func.count()).select_from(action_exports)),
            "outcomes": connection.scalar(select(func.count()).select_from(action_outcomes)),
        }
    source = _source(version_id, analysis_run_id, facts)

    created = service.create_draft(source, facts, "create-a")
    replay = service.create_draft(source, facts, "create-a")
    with pytest.raises(ActionTransitionInvalid):
        service.approve(created.id, created.current_revision, "too early", "approve-0")
    reviewed = service.review(
        created.id,
        created.current_revision,
        "Evidence reviewed",
        "review-a",
    )
    adjusted = service.adjust(
        reviewed.id,
        reviewed.current_revision,
        ActionAdjustment(quantity=Decimal("48"), budget_brl=Decimal("960.00")),
        "Supplier MOQ changed",
        "adjust-a",
    )
    approved = service.approve(
        adjusted.id,
        adjusted.current_revision,
        "Approved for Demo export",
        "approve-a",
    )
    exported = service.export(approved.id, approved.current_revision, "export-a")
    second_export = service.export(
        approved.id,
        approved.current_revision,
        "export-b",
    )
    with pytest.raises(ActionInvalid):
        service.record_outcome(
            approved.id,
            approved.current_revision,
            review_date=date(2026, 8, 31),
            synthetic_result={"units_received": "48"},
            evidence=(replace(facts[0], value="999999"), *facts[1:]),
            conclusion="achieved",
            reason="Synthetic result met the threshold",
            idempotency_key="outcome-invalid-evidence",
        )
    with migrated_engine.connect() as connection:
        snapshot_key = connection.scalar(
            select(storage_objects.c.object_key)
            .select_from(
                analysis_artifacts.join(
                    storage_objects,
                    analysis_artifacts.c.storage_object_id == storage_objects.c.id,
                )
            )
            .where(analysis_artifacts.c.run_id == analysis_run_id)
        )
    assert snapshot_key is not None
    original_snapshot = storage.objects[snapshot_key].content
    storage.objects[snapshot_key].content = b"{}"
    with pytest.raises(ActionInvalid):
        service.record_outcome(
            approved.id,
            approved.current_revision,
            review_date=date(2026, 8, 31),
            synthetic_result={"units_received": "48"},
            evidence=facts,
            conclusion="inconclusive",
            reason="Authority must be revalidated",
            idempotency_key="outcome-stale-source",
        )
    storage.objects[snapshot_key].content = original_snapshot
    with pytest.raises(ActionInvalid):
        service.record_outcome(
            approved.id,
            approved.current_revision,
            review_date=date(2026, 8, 31),
            synthetic_result={"customer_ids": "CUST-123"},
            evidence=facts,
            conclusion="inconclusive",
            reason="Unsafe non-synthetic identifier",
            idempotency_key="outcome-invalid-identifier-list",
        )
    outcome = service.record_outcome(
        approved.id,
        approved.current_revision,
        review_date=date(2026, 8, 31),
        synthetic_result={"units_received": "48", "stockout_days": "0"},
        evidence=facts,
        conclusion="achieved",
        reason="Synthetic result met the threshold",
        idempotency_key="outcome-a",
    )

    current = service.get(approved.id)
    create_replay_after_terminal = service.create_draft(source, facts, "create-a")
    review_replay_after_terminal = service.review(
        created.id,
        created.current_revision,
        "Evidence reviewed",
        "review-a",
    )
    adjust_replay_after_terminal = service.adjust(
        reviewed.id,
        reviewed.current_revision,
        ActionAdjustment(quantity=Decimal("48"), budget_brl=Decimal("960.00")),
        "Supplier MOQ changed",
        "adjust-a",
    )
    approve_replay_after_children = service.approve(
        adjusted.id,
        adjusted.current_revision,
        "Approved for Demo export",
        "approve-a",
    )
    assert replay == created
    assert create_replay_after_terminal == created
    assert review_replay_after_terminal == reviewed
    assert adjust_replay_after_terminal == adjusted
    assert approve_replay_after_children == approved
    assert created.status == "new" and created.current_revision == 1
    assert reviewed.status == "reviewed"
    assert adjusted.status == "reviewed" and adjusted.current_revision == 2
    assert adjusted.revisions[-1].quantity == Decimal("48")
    assert approved.status == "approved"
    assert tuple(item.decision_ordinal for item in approved.decisions) == (1, 2, 3)
    assert tuple(item.command for item in approved.decisions) == (
        "review",
        "adjust",
        "approve",
    )
    assert exported.status == "available"
    assert exported.note == "Not sent to an external platform"
    assert exported.storage_object_id is not None
    assert second_export.id != exported.id
    assert second_export.storage_object_id == exported.storage_object_id
    assert service.open_export(approved.id, exported.id).startswith(b"PK")
    assert service.open_export(approved.id, second_export.id).startswith(b"PK")
    workbook = load_workbook(
        BytesIO(service.open_export(approved.id, exported.id)),
        data_only=False,
        read_only=True,
    )
    export_values = {
        row[0].value: row[1].value
        for row in workbook.active.iter_rows(min_row=4, max_col=2)
    }
    assert export_values["Decision"] == "approve"
    assert export_values["Decision reason"] == "Approved for Demo export"
    assert current.status == "approved"
    assert not hasattr(current, "executed")
    assert outcome.outcome_revision == 1
    assert current.outcomes == (outcome,)
    assert any(item.content.startswith(b"PK") for item in storage.objects.values())

    with migrated_engine.connect() as connection:
        counts = {
            "cards": connection.scalar(select(func.count()).select_from(action_cards)),
            "revisions": connection.scalar(
                select(func.count()).select_from(action_card_revisions)
            ),
            "decisions": connection.scalar(
                select(func.count()).select_from(action_decisions)
            ),
            "exports": connection.scalar(select(func.count()).select_from(action_exports)),
            "outcomes": connection.scalar(select(func.count()).select_from(action_outcomes)),
        }
    assert counts == {
        "cards": baseline_counts["cards"] + 1,
        "revisions": baseline_counts["revisions"] + 2,
        "decisions": baseline_counts["decisions"] + 3,
        "exports": baseline_counts["exports"] + 2,
        "outcomes": baseline_counts["outcomes"] + 1,
    }


def test_export_recovers_promotion_ack_loss(
    migrated_engine: Engine,
) -> None:
    promotion_storage = PromoteAcknowledgementLostStorage()
    service, _storage, version_id, analysis_run_id, facts = _service(
        migrated_engine,
        promotion_storage,
    )
    created = service.create_draft(
        _source(version_id, analysis_run_id, facts),
        facts,
        "promotion-create",
    )
    reviewed = service.review(created.id, 1, "Reviewed", "promotion-review")
    approved = service.approve(reviewed.id, 1, "Approved", "promotion-approve")
    promotion_storage.failures_remaining = 2

    exported = service.export(approved.id, 1, "promotion-export")

    assert promotion_storage.failures_remaining == 0
    assert service.open_export(approved.id, exported.id).startswith(b"PK")
    assert not any("/staging/" in key for key in promotion_storage.objects)


def test_export_tracks_staging_cleanup_failure(
    migrated_engine: Engine,
) -> None:
    cleanup_storage = StagingDeleteFailureStorage()
    cleanup_service, _storage, version_id, analysis_run_id, facts = _service(
        migrated_engine,
        cleanup_storage,
    )
    created = cleanup_service.create_draft(
        _source(version_id, analysis_run_id, facts),
        facts,
        "cleanup-create",
    )
    reviewed = cleanup_service.review(created.id, 1, "Reviewed", "cleanup-review")
    approved = cleanup_service.approve(reviewed.id, 1, "Approved", "cleanup-approve")
    cleanup_storage.armed = True

    cleanup_service.export(approved.id, 1, "cleanup-export")

    with migrated_engine.connect() as connection:
        cleanup_row = connection.execute(
            select(
                storage_objects.c.state,
                storage_objects.c.expires_at,
            ).where(
                storage_objects.c.workspace_id == WORKSPACE_ID,
                storage_objects.c.purpose == "temporary_upload",
                storage_objects.c.object_key.like("%/staging/%"),
            )
        ).one()
    assert cleanup_row.state == "quarantined"
    assert cleanup_row.expires_at == NOW


def test_export_cleans_unledgered_staging_after_database_failure(
    migrated_engine: Engine,
) -> None:
    service, storage, version_id, analysis_run_id, facts = _service(migrated_engine)
    created = service.create_draft(
        _source(version_id, analysis_run_id, facts),
        facts,
        "ledger-create",
    )
    reviewed = service.review(created.id, 1, "Reviewed", "ledger-review")
    approved = service.approve(reviewed.id, 1, "Approved", "ledger-approve")
    before = set(storage.objects)
    FirstCommitFailsUnitOfWork.commit_calls = 0
    failing = ActionService(
        migrated_engine,
        storage,
        WORKSPACE_ID,
        clock=lambda: NOW,
        uow_factory=FirstCommitFailsUnitOfWork,
    )

    with pytest.raises(RuntimeError, match="injected_action_ledger_commit_failure"):
        failing.export(approved.id, 1, "ledger-export")

    assert set(storage.objects) == before
    with migrated_engine.connect() as connection:
        assert connection.scalar(
            select(func.count()).select_from(action_exports).where(
                action_exports.c.action_id == approved.id
            )
        ) == 0


def test_unknown_promotion_is_durable_and_retryable_after_lifecycle_cleanup(
    migrated_engine: Engine,
) -> None:
    storage = UnrecoverablePromotionStorage()
    service, _storage, version_id, analysis_run_id, facts = _service(
        migrated_engine,
        storage,
    )
    created = service.create_draft(
        _source(version_id, analysis_run_id, facts),
        facts,
        "unknown-promotion-create",
    )
    reviewed = service.review(created.id, 1, "Reviewed", "unknown-promotion-review")
    approved = service.approve(
        reviewed.id,
        1,
        "Approved",
        "unknown-promotion-approve",
    )
    storage.armed = True

    with pytest.raises(RuntimeError, match="injected_unrecoverable_promotion"):
        service.export(approved.id, 1, "unknown-promotion-export")

    with migrated_engine.connect() as connection:
        pending = connection.execute(
            select(
                storage_objects.c.object_key,
                storage_objects.c.state,
                storage_objects.c.etag,
            ).where(
                storage_objects.c.workspace_id == WORKSPACE_ID,
                storage_objects.c.purpose == "temporary_upload",
                storage_objects.c.state == "staging",
                storage_objects.c.object_key.like("%/exports/%"),
            )
        ).one()
    assert pending.state == "staging"
    assert pending.etag is None
    assert pending.object_key in storage.objects

    storage.armed = False
    lifecycle = StorageLifecycle(
        migrated_engine,
        storage,
        WORKSPACE_ID,
        clock=lambda: NOW + timedelta(minutes=6),
    )
    assert lifecycle.expire(NOW + timedelta(minutes=6)) == 1
    assert pending.object_key not in storage.objects

    exported = service.export(approved.id, 1, "unknown-promotion-export")
    assert service.open_export(approved.id, exported.id).startswith(b"PK")


def test_revision_and_idempotency_authority_fail_closed(migrated_engine: Engine) -> None:
    service, storage, version_id, analysis_run_id, facts = _service(migrated_engine)
    source = _source(version_id, analysis_run_id, facts)
    created = service.create_draft(
        source,
        facts,
        "create-a",
    )

    with pytest.raises(ActionIdempotencyConflict):
        service.create_draft(
            _source(
                version_id,
                analysis_run_id,
                facts,
                suggestion="Different request",
            ),
            facts,
            "create-a",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(
                _source(version_id, analysis_run_id, facts),
                quantity=Decimal("999999"),
            ),
            facts,
            "invalid-action-number",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            _source(version_id, analysis_run_id, facts),
            (replace(facts[0], value="999999"), *facts[1:]),
            "invalid-fact-number",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            _source(version_id, analysis_run_id, facts),
            (
                FactRef(
                    alias=facts[0].alias,
                    evidence_state="unknown",
                    source_ref=facts[0].source_ref,
                    value=None,
                ),
            ),
            "invalid-evidence",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(
                source,
                chat_turn_id=uuid4(),
                chat_tool="synthetic_chat_tool",
                answer_version="answer-v1",
            ),
            facts,
            "mixed-source-authority",
        )
    snapshot = AnalysisService(
        migrated_engine,
        storage,
        WORKSPACE_ID,
        clock=lambda: NOW,
    ).get_snapshot(analysis_run_id)
    other = next(
        item
        for item in snapshot["result"]["items"]
        if item["sku_id"] != source.target
    )
    evidence_alias = facts[0].alias.split("|", 1)[0]
    mismatched = replace(
        facts[0],
        alias=(
            f"{evidence_alias}|result.items.{other['sku_id']}.recommended_quantity"
        ),
        source_ref=(
            f"analysis:{analysis_run_id}:{evidence_alias}|"
            f"result.items.{other['sku_id']}.recommended_quantity"
        ),
        value=str(other["recommended_quantity"]),
    )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(source, quantity=Decimal(mismatched.value)),
            (mismatched, *facts[1:]),
            "cross-bound-analysis-fact",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(source, target="SYNTH-SKU-00"),
            facts,
            "prefix-collision-target",
        )
    reviewed = service.review(created.id, 1, "Reviewed", "review-a")
    adjusted = service.adjust(
        reviewed.id,
        1,
        ActionAdjustment(quantity=Decimal("48")),
        "first adjustment",
        "adjust-a",
    )
    with pytest.raises(ActionRevisionConflict):
        service.adjust(
            adjusted.id,
            1,
            ActionAdjustment(quantity=Decimal("50")),
            "stale",
            "adjust-stale",
        )


def test_public_session_does_not_gain_action_approved_after_session_start(
    migrated_engine: Engine,
) -> None:
    service, storage, version_id, analysis_run_id, facts = _service(migrated_engine)
    created = service.create_draft(
        _source(version_id, analysis_run_id, facts),
        facts,
        "create-before-session",
    )
    reviewed = service.review(created.id, 1, "Reviewed", "review-before-session")
    session_started_at = NOW + timedelta(seconds=1)
    later = ActionService(
        migrated_engine,
        storage,
        WORKSPACE_ID,
        clock=lambda: NOW + timedelta(seconds=2),
    )
    later.approve(reviewed.id, 1, "Approved later", "approve-after-session")

    visible_ids = {
        item.id
        for item in later.list_public(
            version_id,
            session_started_at,
            MAIN_ACTION_SCOPE,
        )
    }
    assert reviewed.id not in visible_ids


def test_public_session_timestamp_tie_is_fail_closed(
    migrated_engine: Engine,
) -> None:
    service, _storage, version_id, analysis_run_id, facts = _service(migrated_engine)
    created = service.create_draft(
        _source(version_id, analysis_run_id, facts),
        facts,
        "create-timestamp-tie",
    )
    reviewed = service.review(created.id, 1, "Reviewed", "review-timestamp-tie")
    approved = service.approve(
        reviewed.id,
        1,
        "Approved at the same timestamp",
        "approve-timestamp-tie",
    )

    assert service.list_public(version_id, NOW, MAIN_ACTION_SCOPE) == ()
    assert approved.id in {
        item.id
        for item in service.list_public(
            version_id,
            NOW + timedelta(seconds=1),
            MAIN_ACTION_SCOPE,
        )
    }


def test_public_action_projects_simulation_inputs_from_its_completed_run(
    migrated_engine: Engine,
) -> None:
    service, _storage, version_id, analysis_run_id, facts = _service(migrated_engine)
    source = _source(version_id, analysis_run_id, facts)
    assert source.budget_brl is not None
    created = service.create_draft(source, facts, "create-simulation-inputs")
    reviewed = service.review(created.id, 1, "Reviewed", "review-simulation-inputs")
    approved = service.approve(reviewed.id, 1, "Approved", "approve-simulation-inputs")

    public = service.list_public(
        version_id,
        NOW + timedelta(seconds=1),
        MAIN_ACTION_SCOPE,
    )

    projected = next(item for item in public if item.id == approved.id)
    assert projected.simulation_inputs is not None
    assert projected.simulation_inputs.unit_cost_brl is not None
    assert projected.simulation_inputs.precomputed_daily_velocity is not None
    assert projected.simulation_inputs.baseline_budget_brl == source.budget_brl
    assert projected.simulation_inputs.currency == "BRL"


def test_source_and_adjustment_reject_real_data_patterns(
    migrated_engine: Engine,
) -> None:
    service, _storage, version_id, analysis_run_id, facts = _service(migrated_engine)
    source = _source(version_id, analysis_run_id, facts)
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(
                source,
                scope={
                    **source.scope,
                    "note": "contact person@example.test",
                },
            ),
            facts,
            "unsafe-source",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(
                source,
                expected_impact={"customer_id": "CUST-123"},
            ),
            facts,
            "unsafe-nested-source",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(
                source,
                expected_impact={"customer_ids": "CUST-123"},
            ),
            facts,
            "unsafe-plural-source",
        )
    with pytest.raises(ActionInvalid):
        service.create_draft(
            replace(
                source,
                scope={**source.scope, "customer_ids": ["CUST-123"]},
            ),
            facts,
            "unsafe-plural-list-source",
        )

    created = service.create_draft(source, facts, "safe-source")
    reviewed = service.review(created.id, 1, "Reviewed", "safe-review")
    with pytest.raises(ActionInvalid):
        service.adjust(
            reviewed.id,
            1,
            ActionAdjustment(suggestion="Contact person@example.test"),
            "unsafe adjustment",
            "unsafe-adjustment",
        )
    with pytest.raises(ActionInvalid):
        service.adjust(
            reviewed.id,
            1,
            ActionAdjustment(expected_impact={"customer_id": "CUST-123"}),
            "unsafe nested adjustment",
            "unsafe-nested-adjustment",
        )


def test_forecast_and_profit_bridge_are_exact_action_authorities(
    migrated_engine: Engine,
) -> None:
    actions, storage, version_id, _analysis_run_id, _analysis_facts = _service(
        migrated_engine
    )
    forecasts = ForecastService(migrated_engine, storage, WORKSPACE_ID, clock=lambda: NOW)
    request = ForecastRequest(
        candidate=ProductCandidate(
            product_name="Synthetic Portable Organizer",
            category="travel_bag",
            attributes=("portable", "zippered", "compact"),
            planned_launch_date=date(2026, 8, 20),
            planned_price_brl=Decimal("119.90"),
            expected_discount_brl=Decimal("5.00"),
            unit_cost_brl=Decimal("42.00"),
            opening_inventory_units=80,
            moq_units=24,
            lead_time_days=18,
            planned_daily_ad_brl=Decimal("12.00"),
        ),
        safety_stock_units=20,
        assumptions=("synthetic_launch_ramp",),
        missing_fields=(),
    )
    draft = forecasts.create(
        version_id,
        request,
        idempotency_key="action-authority-forecast",
    )
    forecasts.confirm_analogs(
        draft.id,
        tuple(item.sku_id for item in draft.analogs[:2]),
    )
    completed = forecasts.run(draft.id)
    assert completed.result is not None
    quantity = completed.result["recommended_first_order_units"]
    launch_date = completed.input_snapshot["candidate"]["planned_launch_date"]
    forecast_facts = (
        FactRef(
            alias="result.recommended_first_order_units",
            evidence_state="derived",
            source_ref=(
                f"forecast:{completed.id}:result.recommended_first_order_units"
            ),
            value=str(quantity),
        ),
        FactRef(
            alias="input_snapshot.candidate.planned_launch_date",
            evidence_state="assumed",
            source_ref=(
                f"forecast:{completed.id}:"
                "input_snapshot.candidate.planned_launch_date"
            ),
            value=str(launch_date),
        ),
    )
    forecast_source = ActionSource(
        source_type="new_product_forecast",
        dataset_version_id=version_id,
        suggestion="Prepare the synthetic first order",
        target=f"SYNTH-FORECAST-{completed.id}",
        period_start=date(2026, 8, 20),
        period_end=date(2026, 9, 18),
        scope={
            "store_id": "SYNTH-STORE-01",
            "currency": "BRL",
            "period_start": "2026-08-20",
            "period_end": "2026-09-18",
        },
        quantity=Decimal(str(quantity)),
        budget_brl=None,
        action_date=date.fromisoformat(str(launch_date)),
        threshold=None,
        expected_impact={},
        confidence=completed.confidence,
        limitations=("synthetic_demo_only",),
        analysis_run_id=None,
        forecast_id=completed.id,
        bridge_id=None,
        chat_turn_id=None,
        chat_tool=None,
        answer_version=None,
    )
    forecast_action = actions.create_draft(
        forecast_source,
        forecast_facts,
        "create-from-forecast-authority",
    )
    assert forecast_action.revisions[0].forecast_id == completed.id
    with pytest.raises(ActionInvalid):
        actions.create_draft(
            replace(forecast_source, target="SYNTH-FORECAST-OTHER"),
            forecast_facts,
            "forecast-wrong-target",
        )
    with pytest.raises(ActionInvalid):
        actions.create_draft(
            replace(
                forecast_source,
                scope={**forecast_source.scope, "sku_ids": ["SYNTH-SKU-001"]},
            ),
            forecast_facts,
            "forecast-wrong-scope",
        )
    wrong_confidence = "low" if completed.confidence != "low" else "high"
    with pytest.raises(ActionInvalid):
        actions.create_draft(
            replace(forecast_source, confidence=wrong_confidence),
            forecast_facts,
            "forecast-wrong-confidence",
        )
    with pytest.raises(ActionInvalid):
        actions.create_draft(
            replace(forecast_source, limitations=()),
            forecast_facts,
            "forecast-omitted-limitations",
        )
    with pytest.raises(ActionInvalid):
        actions.create_draft(
            replace(forecast_source, quantity=None, action_date=None),
            (
                FactRef(
                    alias="backtest",
                    evidence_state="derived",
                    source_ref=f"forecast:{completed.id}:backtest",
                    value=None,
                ),
            ),
            "forecast-missing-as-derived",
        )

    bridges = ProfitBridgeService(
        migrated_engine,
        storage,
        WORKSPACE_ID,
        analysis_service=AnalysisService(
            migrated_engine,
            storage,
            WORKSPACE_ID,
            clock=lambda: NOW,
        ),
        clock=lambda: NOW,
    )
    bridge = bridges.run(
        version_id,
        current_period=(date(2026, 7, 1), date(2026, 7, 30)),
        comparison_period=(date(2026, 6, 1), date(2026, 6, 30)),
        scope={"store_id": "SYNTH-STORE-01", "currency": "BRL"},
    )
    item = next(value for value in bridge.items if value.amount_brl is not None)
    bridge_fact = FactRef(
        alias=f"items.{item.driver}.amount_brl",
        evidence_state=item.evidence_state,
        source_ref=(f"profit_bridge:{bridge.id}:items.{item.driver}.amount_brl"),
        value=str(item.amount_brl),
    )
    bridge_action = actions.create_draft(
        ActionSource(
            source_type="profit_bridge",
            dataset_version_id=version_id,
            suggestion="Review the synthetic profit driver",
            target="SYNTH-PROFIT-DRIVER",
            period_start=bridge.current_period[0],
            period_end=bridge.current_period[1],
            scope=bridge.scope,
            quantity=None,
            budget_brl=None,
            action_date=None,
            threshold=None,
            expected_impact={},
            confidence="medium",
            limitations=("synthetic_demo_only",),
            analysis_run_id=None,
            forecast_id=None,
            bridge_id=bridge.id,
            chat_turn_id=None,
            chat_tool=None,
            answer_version=None,
        ),
        (bridge_fact,),
        "create-from-bridge-authority",
    )
    assert bridge_action.revisions[0].bridge_id == bridge.id
