from __future__ import annotations

import json
import socket
from collections import defaultdict
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from src.synthetic.reference_contract import (
    canonical_role_count,
    classify_reference,
)
from src.synthetic.generator import generate_and_write, generate_demo
from src.synthetic.manifest import verify_bundle_directory

REQUIRED_SCENARIOS = {
    "sales_ads_growth",
    "inventory_stockout",
    "fifo_aging",
    "profit_decline_ad_spend",
    "new_product_low_base_high",
    "profit_bridge_residual",
    "chat_clarification",
    "action_outcome",
}
FIXTURE_ROOT = Path(__file__).resolve().parents[1] / "fixtures" / "synthetic" / "v1"


def test_same_seed_produces_same_manifest_and_file_bytes(tmp_path: Path) -> None:
    first = generate_and_write(tmp_path / "a", seed=20260813)
    second = generate_and_write(tmp_path / "b", seed=20260813)

    assert first.manifest_bytes == second.manifest_bytes
    assert first.file_hashes == second.file_hashes
    assert {
        file.relative_path: file.content for file in first.files
    } == {file.relative_path: file.content for file in second.files}


def test_bundle_contains_named_acceptance_scenarios() -> None:
    bundle = generate_demo(seed=20260813)

    assert set(bundle.manifest.scenario_ids) >= REQUIRED_SCENARIOS
    assert bundle.manifest.source_classification == "pure_synthetic"
    assert bundle.manifest.currency == "BRL"
    assert bundle.manifest.date_range == ("2026-03-15", "2026-07-31")
    assert bundle.manifest.reporting_period == ("2026-05-01", "2026-07-31")
    assert any(file.relative_path.endswith(".xlsx") for file in bundle.files)


def test_launch_store_is_stable_low_traffic_and_conserves_store_totals() -> None:
    bundle = generate_demo(seed=20260813)
    artifact = next(
        item for item in bundle.files if item.relative_path == "analysis_bundle.json"
    )
    payload = json.loads(artifact.content)
    tables = payload["tables"]

    assert payload["store_catalog"] == [
        {
            "currency": "BRL",
            "display_name_en": "Brazil Main Store",
            "display_name_zh": "巴西主店",
            "has_data": True,
            "lifecycle": "established",
            "opened_on": "2026-05-01",
            "store_id": "SYNTH-STORE-01",
        },
        {
            "currency": "BRL",
            "display_name_en": "Brazil Launch Store",
            "display_name_zh": "巴西新店",
            "has_data": True,
            "lifecycle": "new",
            "opened_on": "2026-07-08",
            "store_id": "SYNTH-STORE-02",
        },
    ]
    assert json.loads(bundle.manifest_bytes)["store_catalog"] == payload["store_catalog"]

    launch_sales = [
        row for row in tables["daily_sales"]
        if row["store_id"] == "SYNTH-STORE-02"
    ]
    launch_ads = [
        row for row in tables["shopee_advertising"]
        if row["store_id"] == "SYNTH-STORE-02"
    ]
    launch_skus = {"SYNTH-SKU-001", "SYNTH-SKU-003", "SYNTH-SKU-006"}
    assert {row["sku_id"] for row in launch_sales} == launch_skus
    assert {row["sku_id"] for row in launch_ads} == launch_skus
    assert min(row["date"] for row in launch_sales) == "2026-07-08"
    assert len(tables["product_catalog"]) == 6

    units_by_day: dict[str, int] = defaultdict(int)
    for row in launch_sales:
        units_by_day[row["date"]] += int(row["units"])
    assert any(units == 0 for units in units_by_day.values())

    main_ads = [
        row for row in tables["shopee_advertising"]
        if row["store_id"] == "SYNTH-STORE-01"
        and row["sku_id"] in launch_skus
        and row["date"] >= "2026-07-08"
    ]
    for field in ("impressions", "clicks"):
        ratio = sum(int(row[field]) for row in launch_ads) / sum(
            int(row[field]) for row in main_ads
        )
        assert Decimal("0.10") <= Decimal(str(ratio)) <= Decimal("0.20")
    impression_ratios = {
        round(int(row["impressions"]) / int(main["impressions"]), 3)
        for row, main in zip(
            launch_ads,
            [
                candidate
                for row in launch_ads
                for candidate in main_ads
                if candidate["date"] == row["date"]
                and candidate["sku_id"] == row["sku_id"]
            ],
            strict=True,
        )
    }
    assert len(impression_ratios) >= 8

    sales_totals: dict[tuple[str, str, str], int] = defaultdict(int)
    for row in tables["daily_sales"]:
        period = row["date"][:7]
        sales_totals[(row["store_id"], row["sku_id"], period)] += int(row["units"])
    outbound_total = 0
    for row in tables["outbound_event"]:
        key = (row["store_id"], row["sku_id"], row["date"][:7])
        assert int(row["quantity"]) == sales_totals[key]
        outbound_total += int(row["quantity"])
    assert outbound_total == sum(int(row["units"]) for row in tables["daily_sales"])

    id_fields = {
        "daily_sales": "order_id",
        "inventory_movement": "movement_id",
        "inventory_receipt_lot": "lot_id",
        "outbound_event": "outbound_id",
        "refund": "refund_id",
        "settlement": "fee_id",
        "fulfillment_cost": "fulfillment_id",
        "fx_effect": "fx_effect_id",
        "other_variable_cost": "cost_id",
    }
    for role, field in id_fields.items():
        identifiers = [row[field] for row in tables[role]]
        assert len(identifiers) == len(set(identifiers))


def test_checked_in_manifest_matches_generated_fixture() -> None:
    generated = generate_demo(seed=20260813)
    checked_manifest = (FIXTURE_ROOT / "manifest.json").read_bytes()

    assert checked_manifest == generated.manifest_bytes
    parsed = json.loads(checked_manifest)
    assert parsed["seed"] == 20260813
    assert parsed["generator"]["version"] == "1.5.0"
    for file in generated.files:
        assert (FIXTURE_ROOT / file.relative_path).read_bytes() == file.content


def test_analysis_bundle_contains_equal_profit_bridge_periods() -> None:
    bundle = generate_demo(seed=20260813)
    artifact = next(
        item for item in bundle.files if item.relative_path == "analysis_bundle.json"
    )
    tables = json.loads(artifact.content)["tables"]

    assert {row["date"][:7] for row in tables["daily_sales"]} == {
        "2026-05",
        "2026-06",
        "2026-07",
    }
    assert {row["currency"] for row in tables["daily_sales"]} == {"BRL"}
    for role in (
        "settlement",
        "fulfillment_cost",
        "fx_effect",
        "other_variable_cost",
    ):
        assert {
            (row["period_start"], row["period_end"])
            for row in tables[role]
        } >= {
            ("2026-05-01", "2026-05-31"),
            ("2026-06-01", "2026-06-30"),
            ("2026-07-01", "2026-07-31"),
        }
    assert {row["date"] for row in tables["outbound_event"]} >= {
        "2026-05-31",
        "2026-06-30",
        "2026-07-31",
    }


def test_cost_workbook_reuses_canonical_cost_rows() -> None:
    bundle = generate_demo(seed=20260813)
    workbook_file = next(
        item
        for item in bundle.files
        if item.relative_path == "bizpulse_demo_costs.xlsx"
    )
    analysis_file = next(
        item for item in bundle.files if item.relative_path == "analysis_bundle.json"
    )
    product_rows = json.loads(analysis_file.content)["tables"]["product_catalog"]
    expected_cost_rows = [
        {
            "sku_id": row["sku_id"],
            "product_name": row["product_name"],
            "currency": "BRL",
            "valid_from": "2026-05-01",
            "unit_cost_brl": row["unit_cost_brl"],
            "source_classification": row["source_classification"],
        }
        for row in product_rows
    ]

    workbook = load_workbook(
        BytesIO(workbook_file.content),
        read_only=True,
        data_only=True,
    )
    try:
        assert workbook.sheetnames == [
            "sku_costs",
            "inventory_receipts",
            "platform_fees",
        ]
        expected = {
            "sku_costs": expected_cost_rows,
            "inventory_receipts": json.loads(analysis_file.content)["tables"][
                "inventory_receipt_lot"
            ],
            "platform_fees": json.loads(analysis_file.content)["tables"][
                "settlement"
            ],
        }
        for sheet_name, expected_rows in expected.items():
            rows = workbook[sheet_name].iter_rows(values_only=True)
            headers = tuple(next(rows))
            records = [
                {
                    header: "" if value is None else value
                    for header, value in zip(headers, row, strict=True)
                }
                for row in rows
            ]
            assert records == expected_rows
    finally:
        workbook.close()


def test_incompatible_legacy_sales_file_cannot_become_sales_authority() -> None:
    decision = classify_reference("bizpulse_demo_sales.csv")

    assert decision.included is False
    assert decision.reason == "currency_and_sku_namespace_mismatch"
    assert decision.currency_decision == "USD_incompatible"
    assert decision.sku_namespace_decision == "BP_incompatible"
    assert canonical_role_count(generate_demo(), "sales_authority") == 1


def test_generation_and_verification_do_not_open_network(
    tmp_path: Path,
    monkeypatch,
) -> None:
    def fail_socket(*args, **kwargs):
        del args, kwargs
        raise AssertionError("synthetic_generation_must_not_open_network")

    monkeypatch.setattr(socket, "socket", fail_socket)
    target = tmp_path / "offline"
    generate_and_write(target, seed=20260813)

    assert verify_bundle_directory(target) == ()
