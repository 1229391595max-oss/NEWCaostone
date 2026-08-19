from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook

from src.actions.contracts import ActionCard, ActionRevision, FactRef
from src.actions.exports import build_action_xlsx, safe_cell


def _approved_card() -> ActionCard:
    created_at = datetime(2026, 8, 14, 15, tzinfo=UTC)
    revision = ActionRevision(
        revision=1,
        suggestion="Reorder 40 units",
        target="SYNTH-SKU-001",
        period_start=date(2026, 7, 1),
        period_end=date(2026, 7, 31),
        scope={"currency": "BRL", "store_id": "SYNTH-STORE-01"},
        quantity=Decimal("40"),
        budget_brl=Decimal("800.00"),
        action_date=date(2026, 8, 20),
        threshold=Decimal("2"),
        expected_impact={"stockout_days": "0"},
        confidence="high",
        limitations=("Synthetic Demo only",),
        facts=(
            FactRef(
                alias="inventory.recommended_quantity",
                evidence_state="derived",
                source_ref="analysis:synthetic",
                value="40",
            ),
        ),
        analysis_run_id=None,
        forecast_id=None,
        bridge_id=None,
        chat_turn_id=None,
        chat_tool=None,
        answer_version=None,
        created_at=created_at,
    )
    return ActionCard(
        id=UUID("11111111-1111-4111-8111-111111111111"),
        workspace_id="synthetic-demo",
        dataset_version_id=UUID("22222222-2222-4222-8222-222222222222"),
        source_type="deterministic_rule",
        status="approved",
        current_revision=1,
        revisions=(revision,),
        created_at=created_at,
        updated_at=created_at,
    )


def test_formula_like_values_are_inert_in_demo_export_cells() -> None:
    assert safe_cell("=HYPERLINK(\"https://evil.example\")") == (
        "'=HYPERLINK(\"https://evil.example\")"
    )
    assert safe_cell("+1+1") == "'+1+1"
    assert safe_cell("-1+1") == "'-1+1"
    assert safe_cell("@SUM(A1:A2)") == "'@SUM(A1:A2)"
    assert safe_cell("ordinary") == "ordinary"


def test_prefixed_value_round_trips_as_text() -> None:
    import xlsxwriter

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, safe_cell("=1+1"))
    workbook.close()
    loaded = load_workbook(BytesIO(output.getvalue()), data_only=False, read_only=True)
    assert loaded.active["A1"].data_type == "s"
    assert loaded.active["A1"].value == "'=1+1"


def test_action_export_metadata_is_pinned_to_action_authority() -> None:
    content = build_action_xlsx(_approved_card())

    with ZipFile(BytesIO(content)) as package:
        core_properties = package.read("docProps/core.xml").decode("utf-8")

    assert "2026-08-14T15:00:00Z" in core_properties
